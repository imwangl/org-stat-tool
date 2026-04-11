import os
import zipfile
import shutil
import openpyxl
import xlrd
from flask import Flask, render_template, request, send_file, jsonify
from collections import defaultdict
from datetime import datetime, timedelta
import uuid

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = '/tmp/org_stat_uploads'
app.config['OUTPUT_FOLDER'] = '/tmp/org_stat_outputs'

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)


def find_file(data_dir, keyword):
    """模糊匹配文件名"""
    for f in os.listdir(data_dir):
        if f.endswith(('.xlsx', '.xls')) and keyword in f:
            return f"{data_dir}/{f}"
    return None

def process_org_stat(data_dir, output_dir, user_time):
    """执行机构数统计"""
    
    # 动态查找文件
    template_file = find_file(data_dir, "模板")
    template2_file = find_file(data_dir, "模版2")
    corp_file = find_file(data_dir, "企业用户数")
    org_file = find_file(data_dir, "机构用户")
    precharge_file = find_file(data_dir, "预充值")
    history_file = find_file(data_dir, "历史客户")
    all_history_file = find_file(data_dir, "全部历史客户")
    monthly_file = find_file(data_dir, "月结用户")
    refund_file = find_file(data_dir, "回款数据")
    market_file = find_file(data_dir, "客户数据汇总表-市场部")
    strategy_file = find_file(data_dir, "客户数据汇总表-战略增长中心")
    
    if not template_file:
        raise ValueError("错误: 未找到模板文件")
    if not org_file:
        raise ValueError("错误: 未找到机构用户文件")
    
    # 复制模板
    template_path = f"{output_dir}/模板.xlsx"
    shutil.copy(template_file, template_path)
    
    # 修改A1单元格标题，添加时间
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    ws.cell(1, 1).value = f"客户数据汇总表（{user_time}）未与财务核对版"
    wb.save(template_path)
    
    # 步骤1-2: 财务表数据 → 模板列1-5,12（列5填充时间周期）
    finance_data = []
    files_to_process = [f for f in [market_file, strategy_file] if f]
    for f in files_to_process:
        try:
            wb = xlrd.open_workbook(f)
            ws = wb.sheet_by_index(0)
            for row in range(3, ws.nrows):
                name = ws.cell_value(row, 3)
                if name:
                    finance_data.append({
                        '部门': ws.cell_value(row, 1),
                        '组别': ws.cell_value(row, 2),
                        '姓名': name,
                        '发生月份': user_time
                    })
        except Exception as e:
            print(f"读取{f}失败: {e}")
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    from openpyxl.styles import Alignment
    
    for i, data in enumerate(finance_data):
        row_num = 4 + i
        ws.cell(row_num, 1).value = i + 1
        ws.cell(row_num, 2).value = data['部门']
        ws.cell(row_num, 3).value = data['组别']
        ws.cell(row_num, 4).value = data['姓名']
        ws.cell(row_num, 5).value = data['发生月份']
        ws.cell(row_num, 12).value = data['发生月份']
        ws.row_dimensions[row_num].height = 15
    
    for row in range(4, 4 + len(finance_data)):
        for col in range(1, 26):
            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
    
    wb.save(template_path)
    
    # ========== 步骤2: 企业用户数列1=姓名, 列4=企业客户数 → 模板列6 ==========
    corp_data = {}
    wb = openpyxl.load_workbook(corp_file)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value  # 列1是用户名称
        count = ws.cell(row, 4).value  # 列4是企业客户数
        if name and count is not None:
            corp_data[str(name).strip()] = int(count)
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    for row in range(4, 4 + len(finance_data)):
        name = ws.cell(row, 4).value
        if name and str(name).strip() in corp_data:
            val = corp_data[str(name).strip()]
            ws.cell(row, 6).value = int(val)
        else:
            ws.cell(row, 6).value = 0
        ws.cell(row, 6).number_format = '0'
    
    wb.save(template_path)
    
    # ========== 步骤3: 机构用户列9 → 列17（空值填1,2,3...）==========
    result_a = f"{output_dir}/机构数明细整理.xlsx"
    shutil.copy(org_file, result_a)
    
    wb = openpyxl.load_workbook(result_a)
    ws = wb.active
    
    # 列17 = 列9（空值填1,2,3...）
    counter = 1
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 9).value  # 列9
        if val is None or str(val).strip() == '':
            ws.cell(row, 17).value = counter
            counter += 1
        else:
            ws.cell(row, 17).value = val
    
    wb.save(result_a)
    print(f"步骤3完成: {ws.max_row}行")
    
    # ========== 步骤4-6: 历史客户/月结用户/预充值 → 列18-20 ==========
    print("\n=== 步骤4-6: 历史客户/月结用户/预充值 → 列18-20 ===")
    
    # 加载预充值、月结用户和历史客户文件
    wb_p = openpyxl.load_workbook(precharge_file)
    wb_h = openpyxl.load_workbook(history_file)
    monthly_file = find_file(data_dir, "月结用户")
    wb_m = openpyxl.load_workbook(monthly_file) if monthly_file else None
    
    wb = openpyxl.load_workbook(result_a)
    ws = wb.active
    
    # 列18: 预充值列8 → 机构用户列18
    ws_p = wb_p.active
    for i in range(2, min(ws_p.max_row + 1, ws.max_row + 1)):
        ws.cell(i, 18).value = ws_p.cell(i, 8).value  # 列8
    
    # 列19: 历史客户列5 → 机构用户列19
    ws_h = wb_h.active
    for i in range(2, min(ws_h.max_row + 1, ws.max_row + 1)):
        ws.cell(i, 19).value = ws_h.cell(i, 5).value  # 列5
    
    # 列20: 月结用户列8 → 机构用户列20
    if wb_m:
        ws_m = wb_m.active
        for i in range(2, min(ws_m.max_row + 1, ws.max_row + 1)):
            ws.cell(i, 20).value = ws_m.cell(i, 8).value  # 列8
    
    wb.save(result_a)
    print("步骤4-6完成")
    
    # ========== 步骤7: 列17在17-20中有重复→删除行 ==========
    print("\n=== 步骤7: 去重（列17在17-20中重复） ===")
    result_b = f"{output_dir}/去重后.xlsx"
    shutil.copy(result_a, result_b)
    
    wb = openpyxl.load_workbook(result_b)
    ws = wb.active
    
    col_values = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        for col in [17, 18, 19, 20]:  # 使用17-20四列
            val = ws.cell(row, col).value
            if val and str(val).strip():
                col_values[str(val).strip()].append(row)
    
    rows_to_delete = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 17).value
        if val and str(val).strip() and len(col_values[str(val).strip()]) > 1:
            rows_to_delete.add(row)
    
    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row)
    
    wb.save(result_b)
    print(f"步骤7完成: 删除{len(rows_to_delete)}行，剩余{ws.max_row}行")
    
    # ========== 步骤8: 回款数据列9手机号→删除行 ==========
    print("\n=== 步骤8: 回款数据去重 ===")
    result_c = f"{output_dir}/最终明细.xlsx"
    shutil.copy(result_b, result_c)
    
    # 读取回款数据手机号（列9）
    refund_file = find_file(data_dir, "回款数据")
    refund_phones = set()
    if refund_file:
        wb_refund = xlrd.open_workbook(refund_file)
        ws_refund = wb_refund.sheet_by_index(0)
        for row in range(ws_refund.nrows):
            val = ws_refund.cell_value(row, 8)  # 列9（索引8）
            if val and str(val).strip():
                refund_phones.add(str(val).strip())
        print(f"回款数据: {len(refund_phones)}个手机号")
    
    # 删除列17值包含回款手机号的行
    wb = openpyxl.load_workbook(result_c)
    ws = wb.active
    
    rows_to_delete = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 17).value
        if val:
            val_str = str(val).strip()
            for phone in refund_phones:
                if phone in val_str:
                    rows_to_delete.add(row)
                    break
    
    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row)
    
    wb.save(result_c)
    print(f"步骤8完成: 删除{len(rows_to_delete)}行，剩余{ws.max_row}行")
    
    # ========== 步骤9: 数据透视统计补贴比例 ==========
    print("\n=== 步骤9: 补贴比例统计 ===")
    wb = openpyxl.load_workbook(result_c)
    ws = wb.active
    
    # 假设列5是补贴比例
    stats = defaultdict(lambda: {'-0.5': 0, '-0.1': 0})
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 2).value  # 用户名称列
        ratio = ws.cell(row, 5).value  # 补贴比例列
        if name and ratio:
            r = str(ratio).strip()
            if r in ['-0.5', '-0.50', '-0.5%']:
                stats[str(name).strip()]['-0.5'] += 1
            elif r in ['-0.1', '-0.10', '-0.1%']:
                stats[str(name).strip()]['-0.1'] += 1
    
    # 生成补贴比例统计
    subsidy_wb = openpyxl.Workbook()
    subsidy_ws = subsidy_wb.active
    subsidy_ws.cell(1, 1).value = "用户名称"
    subsidy_ws.cell(1, 2).value = "-0.5计数"
    subsidy_ws.cell(1, 3).value = "-0.1计数"
    
    for i, (name, counts) in enumerate(sorted(stats.items()), start=2):
        subsidy_ws.cell(i, 1).value = name
        subsidy_ws.cell(i, 2).value = counts['-0.5']
        subsidy_ws.cell(i, 3).value = counts['-0.1']
    
    subsidy_path = f"{output_dir}/补贴比例统计.xlsx"
    subsidy_wb.save(subsidy_path)
    print(f"步骤9完成: {len(stats)}个用户")
    
    # ========== 步骤10-11: 补贴比例 → 模板列18,19 ==========
    print("\n=== 步骤10-11: 补贴比例填充模板 ===")
    wb_sub = openpyxl.load_workbook(subsidy_path)
    ws_sub = wb_sub.active
    
    subsidy_50 = {}
    subsidy_10 = {}
    for row in range(2, ws_sub.max_row + 1):
        name = ws_sub.cell(row, 1).value
        if name:
            subsidy_50[str(name).strip()] = ws_sub.cell(row, 2).value or 0
            subsidy_10[str(name).strip()] = ws_sub.cell(row, 3).value or 0
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    for row in range(4, 4 + len(finance_data)):
        name = ws.cell(row, 4).value
        if name:
            ws.cell(row, 18).value = subsidy_50.get(str(name).strip(), 0)
            ws.cell(row, 19).value = subsidy_10.get(str(name).strip(), 0)
    
    wb.save(template_path)
    print("步骤10-11完成")
    
    # ========== 步骤13: 列22=全部历史客户中姓名出现次数 ==========
    print("\n=== 步骤13: 统计全部历史客户调用数 ===")
    all_history_file = find_file(data_dir, "全部历史客户")
    all_history_counts = {}
    if all_history_file:
        wb_ah = openpyxl.load_workbook(all_history_file)
        ws_ah = wb_ah.active
        for row in range(2, ws_ah.max_row + 1):
            name = ws_ah.cell(row, 2).value  # 列2是姓名
            if name:
                all_history_counts[str(name).strip()] = all_history_counts.get(str(name).strip(), 0) + 1
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    for row in range(4, 4 + len(finance_data)):
        name = ws.cell(row, 4).value
        ws.cell(row, 22).value = all_history_counts.get(str(name).strip(), 0)
    
    wb.save(template_path)
    print("步骤13完成")
    
    # 保存主输出
    final_output = f"{output_dir}/未与财务核对版本.xlsx"
    shutil.copy(template_path, final_output)
    
    # 生成与财务核对版
    template2_path = template2_file
    final_verify_path = f"{output_dir}/与财务核对版（{user_time}）.xlsx"
    shutil.copy(template2_path, final_verify_path)
    
    # 修改与财务核对版A1
    wb2 = openpyxl.load_workbook(final_verify_path)
    ws2 = wb2.active
    ws2.cell(1, 1).value = f"客户数据汇总表（{user_time}）与财务核对版"
    wb2.save(final_verify_path)
    
    # 读取源数据
    source_wb = openpyxl.load_workbook(final_output)
    source_ws = source_wb.active
    
    wb = openpyxl.load_workbook(final_verify_path)
    ws = wb.active
    
    # 复制1-6列和12列
    for row in range(4, 4 + len(finance_data)):
        for col in [1, 2, 3, 4, 5, 6, 12]:
            ws.cell(row, col).value = source_ws.cell(row, col).value
    
    # 复制R、S、V列
    for row in range(4, 4 + len(finance_data)):
        ws.cell(row, 20).value = source_ws.cell(row, 18).value
        ws.cell(row, 21).value = source_ws.cell(row, 19).value
        ws.cell(row, 24).value = source_ws.cell(row, 22).value
    
    # 财务表P列匹配
    finance_p_data = {}
    files_to_process = [f for f in [market_file, strategy_file] if f]
    for f in files_to_process:
        wb_f = xlrd.open_workbook(f)
        ws_f = wb_f.sheet_by_index(0)
        for row in range(3, ws_f.nrows):
            name = ws_f.cell_value(row, 3)
            p_val = ws_f.cell_value(row, 15)
            if name:
                finance_p_data[str(name).strip()] = p_val
    
    for row in range(4, 4 + len(finance_data)):
        name = ws.cell(row, 4).value
        if name and str(name).strip() in finance_p_data:
            ws.cell(row, 16).value = finance_p_data[str(name).strip()]
    
    # 财务表R列匹配
    finance_r_data = {}
    files_to_process = [f for f in [market_file, strategy_file] if f]
    for f in files_to_process:
        wb_f = xlrd.open_workbook(f)
        ws_f = wb_f.sheet_by_index(0)
        for row in range(3, ws_f.nrows):
            name = ws_f.cell_value(row, 3)
            r_val = ws_f.cell_value(row, 17)
            if name:
                finance_r_data[str(name).strip()] = r_val
    
    for row in range(4, 4 + len(finance_data)):
        name = ws.cell(row, 4).value
        if name and str(name).strip() in finance_r_data:
            ws.cell(row, 19).value = finance_r_data[str(name).strip()]
    
    # ========== 与财务核对版逻辑 ==========
    # 列O = 列P + 列R + 列X（列15 = 列16 + 列19 + 列24）
    for row in range(4, 4 + len(finance_data)):
        p_val = ws.cell(row, 16).value or 0
        r_val = ws.cell(row, 19).value or 0
        x_val = ws.cell(row, 24).value or 0
        try:
            ws.cell(row, 15).value = float(p_val) + float(r_val) + float(x_val)
        except:
            ws.cell(row, 15).value = 0
    
    # 列R = 列S + 列T（列18 = 列19 + 列20）
    for row in range(4, 4 + len(finance_data)):
        s_val = ws.cell(row, 19).value or 0
        t_val = ws.cell(row, 20).value or 0
        try:
            ws.cell(row, 18).value = float(s_val) + float(t_val)
        except:
            ws.cell(row, 18).value = 0
    
    # 列M = 列O（列13 = 列15）
    for row in range(4, 4 + len(finance_data)):
        o_val = ws.cell(row, 15).value or 0
        ws.cell(row, 13).value = o_val
    
    wb.save(final_verify_path)
    
    return final_output, final_verify_path

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload_single', methods=['POST'])
def upload_single():
    """单文件上传处理 - 11个独立文件上传"""
    user_time = request.form.get('time', '3月')
    
    # 创建任务目录
    task_id = str(uuid.uuid4())
    task_dir = os.path.join(app.config['UPLOAD_FOLDER'], task_id)
    output_dir = os.path.join(app.config['OUTPUT_FOLDER'], task_id)
    os.makedirs(task_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    
    # 文件名映射 (表单字段名 -> 期望的文件名)
    # 注意：process_org_stat 用 find_file 模糊匹配，所以文件名包含关键词即可
    file_map = {
        'org_user': '机构用户.xlsx',        # 机构用户
        'market': '客户数据市场部.xlsx',    # 客户数据汇总表-市场部
        'strategy': '客户数据战略中心.xlsx', # 客户数据汇总表-战略增长中心
        'history': '历史客户.xlsx',         # 历史客户
        'template1': '模板.xlsx',          # 未与财务核对模板
        'template2': '模版2.xlsx',         # 与财务核对模板
        'company': '企业用户数.xlsx',       # 企业用户数
        'all_history': '历史客户_全部.xlsx', # 全部历史客户
        'prepaid': '预充值.xlsx',           # 预充值
        'monthly': '月结用户.xlsx',
        'repayment': '回款数据-需去重.xlsx',         # 月结用户
    }
    
    # 保存上传的文件
    for key, filename in file_map.items():
        f = request.files.get(key)
        if f and f.filename:
            save_name = filename
            f.save(os.path.join(task_dir, save_name))
            print(f"保存文件: {save_name}")
    
    # 处理
    try:
        final_output, final_verify = process_org_stat(task_dir, output_dir, user_time)
        return jsonify({
            'success': True,
            'task_id': task_id,
            'files': [
                {'name': '未与财务核对版本.xlsx', 'path': final_output},
                {'name': '与财务核对版.xlsx', 'path': final_verify},
                {'name': '补贴比例统计.xlsx', 'path': f"{output_dir}/补贴比例统计.xlsx"}
            ]
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/upload', methods=['POST'])
def upload():
    if 'files' not in request.files:
        return jsonify({'error': '请上传Excel文件'}), 400
    
    files = request.files.getlist('files')
    user_time = request.form.get('time', '3月')
    
    if not files or files[0].filename == '':
        return jsonify({'error': '未选择文件'}), 400
    
    # 创建任务目录
    task_id = str(uuid.uuid4())
    task_dir = os.path.join(app.config['UPLOAD_FOLDER'], task_id)
    output_dir = os.path.join(app.config['OUTPUT_FOLDER'], task_id)
    os.makedirs(task_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    
    # 保存上传的文件
    for f in files:
        f.save(os.path.join(task_dir, f.filename))
    
    # 找到数据目录
    data_dir = None
    for root, dirs, files_list in os.walk(task_dir):
        xlsx_files = [f for f in files_list if not f.startswith('._') and (f.endswith('.xlsx') or f.endswith('.xls'))]
        if xlsx_files:
            data_dir = root
            print(f"数据目录: {root}, 文件: {xlsx_files}")
            break
    
    if not data_dir:
        return jsonify({'error': '未找到Excel文件'}), 400
    
    # 处理
    try:
        final_output, final_verify = process_org_stat(data_dir, output_dir, user_time)
        return jsonify({
            'success': True,
            'task_id': task_id,
            'files': [
                {'name': '未与财务核对版本.xlsx', 'path': final_output},
                {'name': '与财务核对版.xlsx', 'path': final_verify},
                {'name': '补贴比例统计.xlsx', 'path': f"{output_dir}/补贴比例统计.xlsx"}
            ]
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/download/<filename>')
def download(filename):
    for root, dirs, files in os.walk(app.config['OUTPUT_FOLDER']):
        if filename in files:
            return send_file(os.path.join(root, filename), as_attachment=True)
    return "文件未找到", 404

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)