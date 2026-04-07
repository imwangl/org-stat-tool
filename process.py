#!/usr/bin/env python3
"""
微风企机构数统计 - 处理脚本
"""

import openpyxl
import glob
import xlrd
import shutil
import re
import os
from datetime import datetime, timedelta
from collections import defaultdict

def find_file(data_dir, keyword):
    """模糊匹配文件名"""
    for f in os.listdir(data_dir):
        if f.endswith(('.xlsx', '.xls')) and keyword in f:
            return f"{data_dir}/{f}"
    return None

def main(data_dir, output_dir, user_time):
    print("=" * 50)
    print("微风企机构数统计")
    print("=" * 50)
    
    os.makedirs(output_dir, exist_ok=True)
    
    # 动态查找文件
    corp_file = find_file(data_dir, "企业用户数")
    org_file = find_file(data_dir, "机构用户")
    precharge_file = find_file(data_dir, "预充值")
    history_file = find_file(data_dir, "历史客户")
    
    template_path = f"{output_dir}/模板.xlsx"
    shutil.copy(f"{data_dir}/模版.xlsx", template_path)
    
    if not all([corp_file, org_file, precharge_file, history_file]):
        print("警告: 部分关键文件未找到!")
        return

    # ========== 步骤1-2: 财务表数据 ==========
    print("\n=== 步骤1-2: 财务表数据 ===")
    finance_data = []
    for f in [f"{data_dir}/客户数据汇总表-市场部26.4.xls", 
              f"{data_dir}/客户数据汇总表-战略增长中心26.4.xls"]:
        try:
            wb = xlrd.open_workbook(f)
            ws = wb.sheet_by_index(0)
            for row in range(3, ws.nrows):
                name = ws.cell_value(row, 3)
                if name:
                    finance_data.append({
                        '部门': ws.cell_value(row, 1),
                        '组别': ws.cell_value(row, 2),
                        '姓名': name,
                        '发生月份': user_time  # 用户提供的调用时间
                    })
        except Exception as e:
            print(f"读取{f}失败: {e}")

    print(f"财务数据: {len(finance_data)}条")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    from openpyxl.styles import Alignment

    # 只填充有数据的行，保留模板其他行
    for i, data in enumerate(finance_data):
        row_num = 4 + i
        ws.cell(row_num, 1).value = i + 1
        ws.cell(row_num, 2).value = data['部门']
        ws.cell(row_num, 3).value = data['组别']
        ws.cell(row_num, 4).value = data['姓名']
        ws.cell(row_num, 5).value = data['发生月份']
        ws.cell(row_num, 12).value = data['发生月份']
        ws.row_dimensions[row_num].height = 15

    # 设置对齐
    for row in range(4, 4 + len(finance_data)):
        for col in range(1, 26):
            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')

    wb.save(template_path)
    print(f"步骤1-2完成: {len(finance_data)}条")

    # ========== 步骤5: 企业用户数 ==========
    print("\n=== 步骤5: 企业用户数 ===")
    corp_data = {}
    wb = openpyxl.load_workbook(corp_file)
    ws = wb.active
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row, 1).value
        count = ws.cell(row, 4).value
        if name and count is not None:
            corp_data[str(name).strip()] = count

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # 填充企业用户数，转换为数字类型
    for row in range(4, 4 + len(finance_data)):
        name = ws.cell(row, 4).value
        if name and str(name).strip() in corp_data:
            val = corp_data[str(name).strip()]
            ws.cell(row, 6).value = int(val)  # 转换为int
        else:
            ws.cell(row, 6).value = 0
        ws.cell(row, 6).number_format = '0'  # 设置数字格式

    wb.save(template_path)
    print("企业用户数填充完成，F列已转为数字类型")

    # ========== 结果a ==========
    print("\n=== 结果a: 步骤6-9 ===")
    result_a = f"{output_dir}/结果a.xlsx"
    shutil.copy(org_file, result_a)

    wb = openpyxl.load_workbook(result_a)
    ws = wb.active
    ws.cell(1, 17).value = "认证企业"
    ws.cell(1, 18).value = "认证企业2"
    ws.cell(1, 19).value = "企业名称"

    for row in range(2, ws.max_row + 1):
        ws.cell(row, 17).value = ws.cell(row, 9).value

    wb_p = openpyxl.load_workbook(precharge_file)
    ws_p = wb_p.active
    for i in range(2, min(ws_p.max_row + 1, ws.max_row + 1)):
        ws.cell(i, 18).value = ws_p.cell(i, 8).value

    wb_h = openpyxl.load_workbook(history_file)
    ws_h = wb_h.active
    for i in range(2, min(ws_h.max_row + 1, ws.max_row + 1)):
        company = ws_h.cell(i, 5).value
        if company:
            ws.cell(i, 19).value = company

    wb.save(result_a)
    print(f"结果a: {ws.max_row}行")

    # ========== 结果b ==========
    print("\n=== 结果b: 步骤10-11 ===")
    result_b = f"{output_dir}/结果b.xlsx"
    shutil.copy(result_a, result_b)

    wb = openpyxl.load_workbook(result_b)
    ws = wb.active

    col_values = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        for col in [17, 18, 19]:
            val = ws.cell(row, col).value
            if val and str(val).strip():
                col_values[str(val).strip()].append(row)

    rows_to_delete = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 17).value
        if val and str(val).strip() and len(col_values[str(val).strip()]) > 1:
            rows_to_delete.add(row)

    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row)

    wb.save(result_b)
    print(f"结果b: {ws.max_row}行")

    # ========== 结果c ==========
    print("\n=== 结果c: 步骤12-13 ===")
    result_c = f"{output_dir}/结果c.xlsx"
    shutil.copy(result_b, result_c)
    wb.save(result_c)
    print(f"结果c: {ws.max_row}行")

    # ========== 步骤15: 补贴比例统计 ==========
    print("\n=== 步骤15: 补贴比例统计 ===")
    wb = openpyxl.load_workbook(result_c)
    ws = wb.active

    stats = defaultdict(lambda: {'-0.5': 0, '-0.1': 0})
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 2).value
        ratio = ws.cell(row, 5).value
        if name and ratio:
            r = str(ratio).strip()
            if r in ['-0.5', '-0.50']:
                stats[str(name).strip()]['-0.5'] += 1
            elif r in ['-0.1', '-0.10']:
                stats[str(name).strip()]['-0.1'] += 1

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "补贴比例统计"
    new_ws.cell(1, 1).value = "用户名称"
    new_ws.cell(1, 2).value = "-0.5计数"
    new_ws.cell(1, 3).value = "-0.1计数"
    for i, (name, counts) in enumerate(sorted(stats.items())):
        new_ws.cell(i+2, 1).value = name
        new_ws.cell(i+2, 2).value = counts['-0.5']
        new_ws.cell(i+2, 3).value = counts['-0.1']

    subsidy_path = f"{output_dir}/补贴比例统计.xlsx"
    new_wb.save(subsidy_path)
    print(f"补贴比例统计: {len(stats)}用户")

    # ========== 步骤16: vlookup填充补贴比例 ==========
    print("\n=== 步骤16: vlookup填充补贴比例 ===")
    wb_sub = openpyxl.load_workbook(subsidy_path)
    ws_sub = wb_sub.active

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    all_names = []
    for row in range(4, 4 + len(finance_data)):
        name = ws.cell(row, 4).value
        if name:
            all_names.append((row, str(name).strip()))

    subsidy_50 = {}
    subsidy_10 = {}
    for row in range(2, ws_sub.max_row + 1):
        name = ws_sub.cell(row, 1).value
        if name:
            subsidy_50[str(name).strip()] = ws_sub.cell(row, 2).value or 0
            subsidy_10[str(name).strip()] = ws_sub.cell(row, 3).value or 0

    for row, name in all_names:
        ws.cell(row, 18).value = subsidy_50.get(name, 0)
        ws.cell(row, 19).value = subsidy_10.get(name, 0)

    print("补贴比例填充完成")

    # ========== 步骤17: 历史客户调用数 ==========
    print("\n=== 步骤17: 历史客户调用数 ===")
    wb_h = openpyxl.load_workbook(history_file)
    ws_h = wb_h.active

    name_counts = {}
    for row in range(2, ws_h.max_row + 1):
        name = ws_h.cell(row, 2).value
        if name:
            name_counts[str(name).strip()] = name_counts.get(str(name).strip(), 0) + 1

    for row, name in all_names:
        ws.cell(row, 22).value = name_counts.get(name, 0)

    wb.save(template_path)
    print("历史客户调用数填充完成")

    # ========== 保存最终输出 ==========
    final_output = f"{output_dir}/2026年X月未与财务核对版本.xlsx"
    shutil.copy(template_path, final_output)

    org_path = f"{output_dir}/机构用户数-整理.xlsx"
    shutil.copy(result_a, org_path)

    print("\n" + "=" * 50)
    print("全部完成!")
    print("=" * 50)
    print(f"输出目录: {output_dir}")
    print(f"主文件: {final_output}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) >= 3:
        data_dir = sys.argv[1]
        output_dir = sys.argv[2]
        user_time = sys.argv[3] if len(sys.argv) > 3 else "3月"
        main(data_dir, output_dir, user_time)
    else:
        print("用法: python process.py <数据目录> <输出目录> <调用时间>")
        print("示例: python process.py ./data ./output \"3月\"")