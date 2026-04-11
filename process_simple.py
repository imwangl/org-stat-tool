#!/usr/bin/env python3
"""
微风企机构数统计 - 简化版逻辑
"""

import openpyxl
import xlrd
import shutil
import re
import os
from collections import defaultdict

def find_file(data_dir, keyword):
    """模糊匹配文件名"""
    for f in os.listdir(data_dir):
        if f.endswith(('.xlsx', '.xls')) and keyword in f:
            return f"{data_dir}/{f}"
    return None

def main(data_dir, output_dir, user_time):
    print("=" * 50)
    print("微风企机构数统计（简化版）")
    print("=" * 50)
    
    os.makedirs(output_dir, exist_ok=True)
    
    # 查找文件
    corp_file = find_file(data_dir, "企业用户数")
    org_file = find_file(data_dir, "机构用户")
    precharge_file = find_file(data_dir, "预充值")
    history_file = find_file(data_dir, "历史客户")
    all_history_file = find_file(data_dir, "全部历史客户")
    monthly_file = find_file(data_dir, "月结用户")
    refund_file = find_file(data_dir, "回款数据")
    
    # 财务表
    market_file = find_file(data_dir, "市场部")
    strategy_file = find_file(data_dir, "战略增长中心")
    
    template_path = f"{output_dir}/模板.xlsx"
    shutil.copy(f"{data_dir}/模板.xlsx", template_path)
    
    print(f"找到文件: 企业用户数={corp_file}, 机构用户={org_file}")
    print(f"预充值={precharge_file}, 历史客户={history_file}, 全部历史客户={all_history_file}, 月结用户={monthly_file}")
    print(f"回款数据={refund_file}, 市场部={market_file}, 战略增长中心={strategy_file}")

    # ========== 步骤1: 财务表数据 → 模板列1-4,12 ==========
    print("\n=== 步骤1: 财务表数据 → 模板 ===")
    finance_data = []
    for f in [market_file, strategy_file]:
        if not f:
            continue
        try:
            wb = xlrd.open_workbook(f)
            ws = wb.sheet_by_index(0)
            for row in range(3, ws.nrows):
                name = ws.cell_value(row, 3)
                if name:
                    finance_data.append({
                        '部门': ws.cell_value(row, 1),
                        '组别': ws.cell_value(row, 2),
                        '姓名': name,
                    })
        except Exception as e:
            print(f"读取{f}失败: {e}")

    print(f"财务数据: {len(finance_data)}条")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    from openpyxl.styles import Alignment

    # 填充列1-4, 12
    for i, data in enumerate(finance_data):
        row_num = 4 + i
        ws.cell(row_num, 1).value = i + 1  # 列1: 序号
        ws.cell(row_num, 2).value = data['部门']  # 列2: 部门
        ws.cell(row_num, 3).value = data['组别']  # 列3: 组别
        ws.cell(row_num, 4).value = data['姓名']  # 列4: 姓名
        ws.cell(row_num, 5).value = user_time  # 列5: 时间周期
        ws.cell(row_num, 12).value = user_time  # 列12: 时间周期
        ws.row_dimensions[row_num].height = 15

    # 设置对齐
    for row in range(4, 4 + len(finance_data)):
        for col in range(1, 26):
            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')

    wb.save(template_path)
    print(f"步骤1完成: {len(finance_data)}条")

    # ========== 步骤2: 企业用户数列4 → 模板列6 ==========
    print("\n=== 步骤2: 企业用户数 → 模板列6 ===")
    corp_data = {}
    wb = openpyxl.load_workbook(corp_file)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value  # 列1是用户名称
        count = ws.cell(row, 4).value  # 列4是企业客户数
        if name and count is not None:
            corp_data[str(name).strip()] = int(count)

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    for row in range(4, 4 + len(finance_data)):
        name = ws.cell(row, 4).value
        if name and str(name).strip() in corp_data:
            ws.cell(row, 6).value = corp_data[str(name).strip()]
        else:
            ws.cell(row, 6).value = 0

    wb.save(template_path)
    print("企业用户数填充完成")

    # ========== 步骤3: 机构用户列9 → 列17（空值填1,2,3...）==========
    print("\n=== 步骤3: 机构用户 → 列17 ===")
    result_a = f"{output_dir}/机构数明细整理.xlsx"
    shutil.copy(org_file, result_a)

    wb = openpyxl.load_workbook(result_a)
    ws = wb.active
    
    # 填充列17，列9→列17（索引8→索引16）
    counter = 1
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 9).value  # 列9是索引9
        if val is None or str(val).strip() == '':
            ws.cell(row, 17).value = counter  # 空值填1,2,3...
            counter += 1
        else:
            ws.cell(row, 17).value = val

    wb.save(result_a)
    print(f"步骤3完成: {ws.max_row}行")

    # ========== 步骤4-6: 历史客户/月结用户/预充值 → 列18-20 ==========
    print("\n=== 步骤4-6: 历史客户/月结用户/预充值 → 列18-20 ===")
    wb = openpyxl.load_workbook(result_a)
    ws = wb.active

    # 列19: 历史客户列5 → 机构用户列19
    wb_h = openpyxl.load_workbook(history_file)
    ws_h = wb_h.active
    for i in range(2, min(ws_h.max_row + 1, ws.max_row + 1)):
        ws.cell(i, 19).value = ws_h.cell(i, 5).value  # 列5

    # 列20: 月结用户列8 → 机构用户列20
    wb_m = openpyxl.load_workbook(monthly_file)
    ws_m = wb_m.active
    for i in range(2, min(ws_m.max_row + 1, ws.max_row + 1)):
        ws.cell(i, 20).value = ws_m.cell(i, 8).value  # 列8

    # 列18: 预充值列8 → 机构用户列18
    wb_p = openpyxl.load_workbook(precharge_file)
    ws_p = wb_p.active
    for i in range(2, min(ws_p.max_row + 1, ws.max_row + 1)):
        ws.cell(i, 18).value = ws_p.cell(i, 8).value  # 列8

    wb.save(result_a)
    print("步骤4-6完成")

    # ========== 步骤7: 列17在17-20中有重复→删除行 ==========
    print("\n=== 步骤7: 去重（列17在17-20中重复） ===")
    result_b = f"{output_dir}/去重后.xlsx"
    shutil.copy(result_a, result_b)

    wb = openpyxl.load_workbook(result_b)
    ws = wb.active

    col_values = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        for col in [17, 18, 19, 20]:
            val = ws.cell(row, col).value
            if val and str(val).strip():
                col_values[str(val).strip()].append(row)

    rows_to_delete = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 17).value
        if val and str(val).strip() and len(col_values[str(val).strip()]) > 1:
            rows_to_delete.add(row)

    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row)

    wb.save(result_b)
    print(f"步骤7完成: 删除{len(rows_to_delete)}行，剩余{ws.max_row}行")

    # ========== 步骤8: 列17包含回款数据手机号→删除行 ==========
    print("\n=== 步骤8: 回款数据去重 ===")
    result_c = f"{output_dir}/最终明细.xlsx"
    shutil.copy(result_b, result_c)

    # 读取回款数据手机号
    refund_phones = set()
    if refund_file:
        wb_refund = xlrd.open_workbook(refund_file)
        ws_refund = wb_refund.sheet_by_index(0)
        for row in range(ws_refund.nrows):
            # 列9是索引8
            val = ws_refund.cell_value(row, 8)  # 列9（备注列，可能有手机号）
            if val and str(val).strip():
                refund_phones.add(str(val).strip())
        print(f"回款数据: {len(refund_phones)}个手机号")

    # 删除列17值包含回款手机号的行
    wb = openpyxl.load_workbook(result_c)
    ws = wb.active

    rows_to_delete = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 17).value
        if val:
            val_str = str(val).strip()
            for phone in refund_phones:
                if phone in val_str:  # 包含就算
                    rows_to_delete.add(row)
                    break

    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row)

    wb.save(result_c)
    print(f"步骤8完成: 删除{len(rows_to_delete)}行，剩余{ws.max_row}行")

    # ========== 步骤9: 数据透视统计补贴比例 ==========
    print("\n=== 步骤9: 补贴比例统计 ===")
    wb = openpyxl.load_workbook(result_c)
    ws = wb.active

    # 假设列5是补贴比例（需要根据实际调整）
    stats = defaultdict(lambda: {'-0.5': 0, '-0.1': 0})
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 2).value  # 用户名称列
        ratio = ws.cell(row, 5).value  # 补贴比例列
        if name and ratio:
            r = str(ratio).strip()
            if r in ['-0.5', '-0.50', '-0.5%']:
                stats[str(name).strip()]['-0.5'] += 1
            elif r in ['-0.1', '-0.10', '-0.1%']:
                stats[str(name).strip()]['-0.1'] += 1

    # 生成补贴比例统计
    subsidy_wb = openpyxl.Workbook()
    subsidy_ws = subsidy_wb.active
    subsidy_ws.cell(1, 1).value = "用户名称"
    subsidy_ws.cell(1, 2).value = "-0.5计数"
    subsidy_ws.cell(1, 3).value = "-0.1计数"
    
    for i, (name, counts) in enumerate(sorted(stats.items()), start=2):
        subsidy_ws.cell(i, 1).value = name
        subsidy_ws.cell(i, 2).value = counts['-0.5']
        subsidy_ws.cell(i, 3).value = counts['-0.1']

    subsidy_path = f"{output_dir}/补贴比例统计.xlsx"
    subsidy_wb.save(subsidy_path)
    print(f"步骤9完成: {len(stats)}个用户")

    # ========== 步骤10-11: 补贴比例 → 模板列18,19 ==========
    print("\n=== 步骤10-11: 补贴比例填充模板 ===")
    wb_sub = openpyxl.load_workbook(subsidy_path)
    ws_sub = wb_sub.active

    subsidy_50 = {}
    subsidy_10 = {}
    for row in range(2, ws_sub.max_row + 1):
        name = ws_sub.cell(row, 1).value
        if name:
            subsidy_50[str(name).strip()] = ws_sub.cell(row, 2).value or 0
            subsidy_10[str(name).strip()] = ws_sub.cell(row, 3).value or 0

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    for row in range(4, 4 + len(finance_data)):
        name = ws.cell(row, 4).value
        if name:
            ws.cell(row, 18).value = subsidy_50.get(str(name).strip(), 0)
            ws.cell(row, 19).value = subsidy_10.get(str(name).strip(), 0)

    wb.save(template_path)
    print("步骤10-11完成")

    # ========== 步骤13: 列22=全部历史客户中姓名出现次数 ==========
    print("\n=== 步骤13: 统计全部历史客户调用数 ===")
    all_history_counts = {}
    if all_history_file:
        wb_ah = openpyxl.load_workbook(all_history_file)
        ws_ah = wb_ah.active
        for row in range(2, ws_ah.max_row + 1):
            name = ws_ah.cell(row, 2).value  # 列2是姓名
            if name:
                all_history_counts[str(name).strip()] = all_history_counts.get(str(name).strip(), 0) + 1

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    for row in range(4, 4 + len(finance_data)):
        name = ws.cell(row, 4).value
        ws.cell(row, 22).value = all_history_counts.get(str(name).strip(), 0)

    wb.save(template_path)
    print("步骤13完成")

    # ========== 步骤12: f77/r77/s77用SUM计算 ==========
    print("\n=== 步骤12: f77/r77/s77用SUM计算 ===")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # 假设f77=列6, r77=列18, s77=列19
    # 需要找到具体单元格位置，根据模板结构调整
    # 这里假设在第77行（索引77）
    ws.cell(77, 6).value = "=SUM(F4:F76)"  # f77
    ws.cell(77, 18).value = "=SUM(R4:R76)"  # r77
    ws.cell(77, 19).value = "=SUM(S4:S76)"  # s77

    wb.save(template_path)
    print("步骤12完成")

    # ========== 保存最终输出 ==========
    final_output = f"{output_dir}/未与财务核对版本.xlsx"
    shutil.copy(template_path, final_output)

    print("\n" + "=" * 50)
    print("全部完成!")
    print("=" * 50)
    print(f"输出目录: {output_dir}")
    print(f"主文件: {final_output}")
    print(f"机构数明细整理: {result_a}")
    print(f"补贴比例统计: {subsidy_path}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) >= 3:
        data_dir = sys.argv[1]
        output_dir = sys.argv[2]
        user_time = sys.argv[3] if len(sys.argv) > 3 else "3月"
        main(data_dir, output_dir, user_time)
    else:
        print("用法: python process.py <数据目录> <输出目录> <调用时间>")
        print("示例: python process.py ./data ./output \"3月\"")