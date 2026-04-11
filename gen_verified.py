#!/usr/bin/env python3
"""
生成与财务核对版
"""

import openpyxl
import xlrd
import shutil
import os

def main():
    data_dir = '/Users/wanglei/.openclaw/workspace/test_org_stat4/3月'
    output_dir = '/Users/wanglei/.openclaw/workspace/机构数统计输出5'
    
    # 1. 打开模版2
    template2_path = f"{data_dir}/模版2.xlsx"
    template2_output = f"{output_dir}/与财务核对版.xlsx"
    shutil.copy(template2_path, template2_output)
    
    wb2 = openpyxl.load_workbook(template2_output)
    ws2 = wb2.active
    
    # 读取未与财务核对版本
    wb1 = openpyxl.load_workbook(f"{output_dir}/未与财务核对版本.xlsx")
    ws1 = wb1.active
    
    # 2. 复制第1-6列和12列
    print("=== 步骤2: 复制1-6列和12列 ===")
    for row in range(4, 72):  # 68条数据
        for col in range(1, 7):
            ws2.cell(row, col).value = ws1.cell(row, col).value
        ws2.cell(row, 12).value = ws1.cell(row, 12).value
    print("步骤2完成")
    
    # 3. 复制R列→T列 (列18→列20)
    print("=== 步骤3: R→T ===")
    for row in range(4, 72):
        ws2.cell(row, 20).value = ws1.cell(row, 18).value
    print("步骤3完成")
    
    # 4. 复制S列→U列 (列19→列21)
    print("=== 步骤4: S→U ===")
    for row in range(4, 72):
        ws2.cell(row, 21).value = ws1.cell(row, 19).value
    print("步骤4完成")
    
    # 5. 复制V列→X列 (列22→列24)
    print("=== 步骤5: V→X ===")
    for row in range(4, 72):
        ws2.cell(row, 24).value = ws1.cell(row, 22).value
    print("步骤5完成")
    
    # 6. P列匹配（根据姓名）
    print("=== 步骤6: P列匹配 ===")
    finance_p_data = {}
    for f in [f"{data_dir}/客户数据汇总表-市场部26.4.xls", 
              f"{data_dir}/客户数据汇总表-战略增长中心26.4.xls"]:
        try:
            wb_f = xlrd.open_workbook(f)
            ws_f = wb_f.sheet_by_index(0)
            for row in range(3, ws_f.nrows):
                name = ws_f.cell_value(row, 3)  # 列D是姓名
                p_val = ws_f.cell_value(row, 15)  # 列P
                if name:
                    finance_p_data[str(name).strip()] = p_val
        except Exception as e:
            print(f"读取{f}失败: {e}")
    
    for row in range(4, 72):
        name = ws2.cell(row, 4).value
        ws2.cell(row, 16).value = finance_p_data.get(str(name).strip(), 0)  # 列P是第16列
    print("步骤6完成")
    
    # 7. R列匹配到S列（根据姓名）
    print("=== 步骤7: R列→S列匹配 ===")
    finance_r_data = {}
    for f in [f"{data_dir}/客户数据汇总表-市场部26.4.xls", 
              f"{data_dir}/客户数据汇总表-战略增长中心26.4.xls"]:
        try:
            wb_f = xlrd.open_workbook(f)
            ws_f = wb_f.sheet_by_index(0)
            for row in range(3, ws_f.nrows):
                name = ws_f.cell_value(row, 3)  # 列D是姓名
                r_val = ws_f.cell_value(row, 17)  # 列R
                if name:
                    finance_r_data[str(name).strip()] = r_val
        except Exception as e:
            print(f"读取{f}失败: {e}")
    
    for row in range(4, 72):
        name = ws2.cell(row, 4).value
        ws2.cell(row, 19).value = finance_r_data.get(str(name).strip(), 0)  # S列是第19列
    print("步骤7完成")
    
    # 8. 列O = 列P + 列R + 列X
    print("=== 步骤8: 计算列O ===")
    for row in range(4, 72):
        p_val = ws2.cell(row, 16).value or 0  # 列P
        r_val = ws2.cell(row, 18).value or 0  # 列R
        x_val = ws2.cell(row, 24).value or 0  # 列X
        ws2.cell(row, 15).value = p_val + r_val + x_val  # 列O = P + R + X
    print("步骤8完成")
    
    # 9. 列M = 列O
    print("=== 步骤9: 计算列M ===")
    for row in range(4, 72):
        o_val = ws2.cell(row, 15).value or 0  # 列O
        ws2.cell(row, 13).value = o_val  # 列M = 列O
    print("步骤9完成")
    
    # 保存
    wb2.save(template2_output)
    print(f"\n与财务核对版生成完成: {template2_output}")

if __name__ == "__main__":
    main()